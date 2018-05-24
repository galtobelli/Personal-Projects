import openpyxl, pprint
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import os

# The next four lines create a sample excel file called test_1.xlsx
df = pd.DataFrame(np.random.randn(15, 12), columns=list('ABCDEFGHIJKL'))
writer = pd.ExcelWriter('test_1.xlsx', engine='openpyxl')
df.to_excel(writer, sheet_name='Sheet1', index=False)
writer.close()

wb = load_workbook('test_1.xlsx')
sheet1 = wb['Sheet1']
my_list = []
for j in range(2, 8):
    my_list.append(sheet1.cell(row = 7, column = j).value)

def HH_C3_gap_col():
    """=IF(Y2=0,"No C3",((R2-P2)/R2))"""
    y,r,p = 1,1,1
    while y < 17 and r < 17 and p < 17:
        for rowNum in range(2, sheet1.max_row + 1):  # skip the first row
            sheet1.cell(row=rowNum, column=14).value = ('=IF(Y' + str(y-sheet1.max_row + 2) +
                                                        '=0,"No C3",((' + 'R' + str(r-sheet1.max_row + 2) + '-' + 'P' + str(p-sheet1.max_row + 2) + ')' + '/' + 'R' + str(r-sheet1.max_row + 2) + '))')
            y+= 1
            r+= 1
            p+= 1
HH_C3_gap_col()

def Average_col():
    """=AVERAGEIFS(AI:AI,AI:AI,"<>No C3",I:I,I2,F:F,F2)"""
    i,f = 1,1
    while i < 17 and f < 17:
        for rowNum in range(2, sheet1.max_row + 1):  # skip the first row
            sheet1.cell(row=rowNum, column=15).value = ('=AVERAGEIFS(AI:AI,AI:AI,"<>No C3",I:I,I' + str(i-sheet1.max_row + 2) + ',F:F,F' + str(f-sheet1.max_row + 2) + ')')
            i+= 1
            f+= 1
Average_col()

def Corrected_C3():
    """=IF(Y2>0,Y2,U2*(1+AJ2))"""
    y,u,aj = 1,1,1
    while y < 17 and u < 17 and aj < 17:
        for rowNum in range(2, sheet1.max_row + 1):  # skip the first row
            sheet1.cell(row=rowNum, column=16).value = ('=IF(Y' + str(y-sheet1.max_row + 2) + '>0,Y' + str(y-sheet1.max_row + 2) + ',U' + str(u-sheet1.max_row + 2) + '*(1+AJ' + str(aj-sheet1.max_row + 2)
            + '))')
            y+= 1
            u+= 1
            aj+= 1
Corrected_C3()

def Corrected_Ad_Index():
    """=IFERROR(IF(AA2=0,AVERAGEIFS(AA:AA,AA:AA,">0",I:I,I2,F:F,F2),AA2),0)"""
    aa,i,f = 1,1,1
    while aa < 17 and i < 17 and f < 17:
        for rowNum in range(2, sheet1.max_row + 1):  # skip the first row
            sheet1.cell(row=rowNum, column=17).value = ('=IFERROR(IF(AA' + str(aa-sheet1.max_row + 2) + '=0,AVERAGEIFS(AA:AA,AA:AA,">0",I:I,I' + str(i-sheet1.max_row + 2)
            + ',F:F,F' + str(f-sheet1.max_row + 2) + '),AA' + str(aa-sheet1.max_row + 2) + '),0)')
            aa+= 1
            i+= 1
            f+= 1
Corrected_Ad_Index()
